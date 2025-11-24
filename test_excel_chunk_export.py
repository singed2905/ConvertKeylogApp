import os
import random
import string
from services.integral.excel_service import ExcelService
import openpyxl

def create_fake_input_excel(file_path, num_rows=5000):
    """Tạo file Excel mẫu với cột int_input và mode."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['STT', 'int_input', 'mode'])
    for i in range(num_rows):
        fake_latex = f"\\int_{{{i}}}^{{{i+1}}} x^{i%5} dx"
        ws.append([i+1, fake_latex, random.choice(['1', '2', '3', '4'])])
    wb.save(file_path)
    wb.close()

def fake_encode_integral(latex, mode):
    """Fake mã hóa LaTeX → keylog"""
    return {
        'success': True,
        'keylog': latex.replace("\\int", "y[") + f"[{mode}]"
    }

def test_excel_chunk_export():
    # Setup
    input_file = "test_input_large.xlsx"
    create_fake_input_excel(input_file, num_rows=15000)  # Tạo file 15.000 dòng

    service = ExcelService()
    ok, rows, err = service.read_excel_file(input_file)
    assert ok, err
    print(f"Read {len(rows)} rows from {input_file}")

    # Fake batch encoding
    batch_results = []
    for latex, mode in rows:
        r = fake_encode_integral(latex, mode)
        batch_results.append({
            'latex': latex,
            'mode': mode,
            'keylog': r['keylog'],
            'status': 'success' if r['success'] else 'error'
        })

    # Export file bằng chunk
    success, output_file, message = service.export_results(batch_results)
    print(message)
    assert success
    assert output_file.endswith('.xlsx')
    assert os.path.exists(output_file)

    # Kiểm tra file xuất có đúng số dòng
    wb = openpyxl.load_workbook(output_file)
    ws = wb.active
    rows_out = list(ws.iter_rows(values_only=True))
    assert len(rows_out) == len(rows) + 1   # +1 dòng header
    assert rows_out[0][-1] == 'keylog'
    print(f"Test xuất file {output_file} PASSED! {len(rows)} dòng.")

if __name__ == "__main__":
    test_excel_chunk_export()
