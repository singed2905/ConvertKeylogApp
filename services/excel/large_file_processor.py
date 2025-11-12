import pandas as pd
import os
import gc
import psutil
from typing import Dict, List, Tuple, Any, Optional, Iterator, Callable
from datetime import datetime
import threading
import time


class LargeFileProcessor:
    """
    OPTIMIZED HIGH-SPEED processor for large Excel files với PRE-VALIDATION
    Features: Single-workbook streaming, 400+ rows/sec, validation trước chunking
    Smart keylog column (strict): only 'keylog'. If absent, create 'keylog'.
    Font styling for keylog column: Flexio Fx799VN, 11pt, bold, black.
    """

    def __init__(self, config: Dict = None):
        self.config = config or {}
        self.processing_cancelled = False
        self.max_memory_mb = 1500
        self.emergency_cleanup = False
        self.max_rows_allowed = 250_000
        self.fast_mode = True

    def _clean_cell_value(self, value):
        """Clean cell value - same logic as ExcelProcessor"""
        if pd.isna(value) or not value:
            return ""

        value_str = str(value).strip()

        if value_str.startswith('[') and value_str.endswith(']'):
            value_str = value_str[1:-1]
            elements = value_str.split(',')
            cleaned_elements = [elem.strip().strip('"').strip("'").strip()
                                for elem in elements if elem.strip()]
            return ','.join(cleaned_elements)

        return value_str

    def get_memory_usage(self) -> float:
        try:
            process = psutil.Process()
            return process.memory_info().rss / 1024 / 1024
        except:
            return 0.0

    def check_memory_limit(self) -> bool:
        return self.get_memory_usage() > self.max_memory_mb

    def emergency_memory_cleanup(self):
        self.emergency_cleanup = True
        gc.collect()

    def _enforce_row_limit(self, total_rows: int):
        if total_rows > self.max_rows_allowed:
            raise Exception(
                f"File có {total_rows:,} dòng, vượt quá giới hạn tối đa {self.max_rows_allowed:,} dòng.\n"
                f"Vui lòng chia nhỏ file hoặc lọc bớt dữ liệu trước khi import."
            )

    def estimate_optimal_chunksize(self, file_path: str) -> int:
        try:
            file_size_mb = os.path.getsize(file_path) / (1024 * 1024)
            if file_size_mb > 100:
                return 3000
            elif file_size_mb > 50:
                return 5000
            elif file_size_mb > 20:
                return 7000
            else:
                return 10000
        except Exception:
            return 5000

    def _get_actual_total_rows(self, file_path: str) -> int:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            actual_total = max(0, (ws.max_row - 1)) if hasattr(ws, 'max_row') and ws.max_row else 0
            wb.close()
            print(f"📊 Actual file dimensions: {actual_total:,} data rows")
            return actual_total
        except Exception as e:
            print(f"⚠️ Could not get actual rows: {e}")
            return 0

    def _detect_keylog_column_strict(self, file_path: str) -> Tuple[bool, str, int]:
        """Strict detection: only 'keylog' (case-insensitive). Returns (has_keylog, 'keylog', index or -1)."""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            wb.close()
            for i, cell in enumerate(header_row):
                if cell is not None and str(cell).strip().lower() == 'keylog':
                    print(f"🔍 Found strict keylog column at position {i + 1}")
                    return True, 'keylog', i
            print("📝 No strict 'keylog' column found. Will create new: 'keylog'")
            return False, 'keylog', -1
        except Exception as e:
            print(f"⚠️ Strict keylog detection failed: {e}")
            return False, 'keylog', -1

    # ========== VALIDATION METHODS - NEW ==========

    def _validate_first_row_before_chunking(self, file_path: str,
                                            shape_a: str, shape_b: str) -> Dict:
        """
        ✅ VALIDATE DÒNG ĐẦU TIÊN TRƯỚC KHI BẮT ĐẦU CHUNKING
        Returns: {'valid': bool, 'issues': List[str], 'warnings': List[str], 'sample_data': Dict}
        """
        result = {
            'valid': True,
            'issues': [],
            'warnings': [],
            'sample_data': {}
        }

        try:
            import openpyxl

            print("🔍 Pre-validation: Checking first row...")

            # Đọc CHỈ header và dòng đầu tiên
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active

            # Header
            header_iter = ws.iter_rows(min_row=1, max_row=1, values_only=True)
            header_row = next(header_iter)
            columns = [str(cell).strip() if cell else "" for cell in header_row]

            # Dòng đầu tiên
            data_iter = ws.iter_rows(min_row=2, max_row=2, values_only=True)
            first_data_row = next(data_iter, None)

            wb.close()

            if not first_data_row:
                result['valid'] = False
                result['issues'].append("❌ File không có dòng dữ liệu")
                print("   ❌ No data rows found")
                return result

            # Kiểm tra cấu trúc columns
            is_valid_structure, missing_cols = self._check_required_columns(
                columns, shape_a, shape_b
            )

            if not is_valid_structure:
                result['valid'] = False
                result['issues'].append(
                    f"❌ Thiếu các cột bắt buộc: {', '.join(missing_cols)}"
                )
                print(f"   ❌ Structure check FAILED: {missing_cols}")
                return result

            print(f"   ✅ Structure check OK")

            # Tạo pandas Series để validate data
            first_row = pd.Series(first_data_row, index=columns)

            # Validate Group A
            data_a = self._extract_shape_data_from_row(first_row, shape_a, 'A')
            has_data_a = any(str(v).strip() for v in data_a.values() if v)

            if not has_data_a:
                result['valid'] = False
                result['issues'].append(f"❌ Dòng 1 - Nhóm A ({shape_a}): Không có dữ liệu")
                print(f"   ❌ Data A check FAILED: No data")
                return result

            # Validate chi tiết dữ liệu A
            issues_a = self._validate_shape_data_detailed(data_a, shape_a, "Nhóm A")
            if issues_a:
                result['valid'] = False
                result['issues'].extend(issues_a)
                print(f"   ❌ Data A check FAILED: {issues_a}")
                return result

            result['sample_data']['group_a'] = data_a
            print(f"   ✅ Data A check OK: {data_a}")

            # Validate Group B nếu có
            if shape_b:
                data_b = self._extract_shape_data_from_row(first_row, shape_b, 'B')
                has_data_b = any(str(v).strip() for v in data_b.values() if v)

                if not has_data_b:
                    result['valid'] = False
                    result['issues'].append(f"❌ Dòng 1 - Nhóm B ({shape_b}): Không có dữ liệu")
                    print(f"   ❌ Data B check FAILED: No data")
                    return result

                issues_b = self._validate_shape_data_detailed(data_b, shape_b, "Nhóm B")
                if issues_b:
                    result['valid'] = False
                    result['issues'].extend(issues_b)
                    print(f"   ❌ Data B check FAILED: {issues_b}")
                    return result

                result['sample_data']['group_b'] = data_b
                print(f"   ✅ Data B check OK: {data_b}")

            print("✅ Pre-validation PASSED - Ready for chunking")

            return result

        except Exception as e:
            result['valid'] = False
            result['issues'].append(f"❌ Lỗi validation: {str(e)}")
            print(f"   ❌ Validation ERROR: {e}")
            return result

    def _check_required_columns(self, columns: List[str], shape_a: str,
                                shape_b: str = None) -> Tuple[bool, List[str]]:
        """Kiểm tra các cột bắt buộc có tồn tại không"""
        missing_columns = []

        # Check Group A
        required_a = self._get_required_columns(shape_a, 'A')
        for col in required_a:
            if col not in columns:
                missing_columns.append(f"Nhóm A - {col}")

        # Check Group B
        if shape_b:
            required_b = self._get_required_columns(shape_b, 'B')
            for col in required_b:
                if col not in columns:
                    missing_columns.append(f"Nhóm B - {col}")

        return len(missing_columns) == 0, missing_columns

    def _extract_shape_data_from_row(self, row: pd.Series, shape_type: str,
                                     group: str) -> Dict:
        """Extract data từ row cho shape cụ thể"""
        data_dict = {}

        if not shape_type:
            return data_dict

        if group == 'A':
            if shape_type == "Điểm":
                data_dict['point_input'] = str(row.get('data_A', '')).strip()
            elif shape_type == "Đường thẳng":
                data_dict['line_A1'] = str(row.get('d_P_data_A', '')).strip()
                data_dict['line_X1'] = str(row.get('d_V_data_A', '')).strip()
            elif shape_type == "Mặt phẳng":
                data_dict['plane_a'] = str(row.get('P1_a', '')).strip()
                data_dict['plane_b'] = str(row.get('P1_b', '')).strip()
                data_dict['plane_c'] = str(row.get('P1_c', '')).strip()
                data_dict['plane_d'] = str(row.get('P1_d', '')).strip()
            elif shape_type == "Đường tròn":
                data_dict['circle_center'] = str(row.get('C_data_I1', '')).strip()
                data_dict['circle_radius'] = str(row.get('C_data_R1', '')).strip()
            elif shape_type == "Mặt cầu":
                data_dict['sphere_center'] = str(row.get('S_data_I1', '')).strip()
                data_dict['sphere_radius'] = str(row.get('S_data_R1', '')).strip()
        else:  # Group B
            if shape_type == "Điểm":
                data_dict['point_input'] = str(row.get('data_B', '')).strip()
            elif shape_type == "Đường thẳng":
                data_dict['line_A2'] = str(row.get('d_P_data_B', '')).strip()
                data_dict['line_X2'] = str(row.get('d_V_data_B', '')).strip()
            elif shape_type == "Mặt phẳng":
                data_dict['plane_a'] = str(row.get('P2_a', '')).strip()
                data_dict['plane_b'] = str(row.get('P2_b', '')).strip()
                data_dict['plane_c'] = str(row.get('P2_c', '')).strip()
                data_dict['plane_d'] = str(row.get('P2_d', '')).strip()
            elif shape_type == "Đường tròn":
                data_dict['circle_center'] = str(row.get('C_data_I2', '')).strip()
                data_dict['circle_radius'] = str(row.get('C_data_R2', '')).strip()
            elif shape_type == "Mặt cầu":
                data_dict['sphere_center'] = str(row.get('S_data_I2', '')).strip()
                data_dict['sphere_radius'] = str(row.get('S_data_R2', '')).strip()

        return data_dict

    def _validate_shape_data_detailed(self, data: Dict, shape: str,
                                      group_name: str) -> List[str]:
        """Validate chi tiết data của 1 shape"""
        issues = []

        if shape == "Điểm":
            point_input = data.get('point_input', '').strip()
            if not point_input:
                issues.append(f"❌ {group_name}: Tọa độ điểm trống")
            else:
                coords = point_input.split(',')
                if len(coords) < 2:
                    issues.append(
                        f"❌ {group_name}: Điểm cần ít nhất 2 tọa độ, "
                        f"nhận được: '{point_input}'"
                    )

        elif shape == "Đường thẳng":
            line_A = data.get('line_A1') or data.get('line_A2', '')
            line_X = data.get('line_X1') or data.get('line_X2', '')

            if not line_A.strip():
                issues.append(f"❌ {group_name}: Điểm trên đường thẳng trống")
            else:
                coords_A = line_A.split(',')
                if len(coords_A) != 3:
                    issues.append(
                        f"❌ {group_name}: Điểm đường thẳng cần 3 tọa độ, "
                        f"nhận được: '{line_A}'"
                    )

            if not line_X.strip():
                issues.append(f"❌ {group_name}: Vector phương trống")
            else:
                coords_X = line_X.split(',')
                if len(coords_X) != 3:
                    issues.append(
                        f"❌ {group_name}: Vector phương cần 3 tọa độ, "
                        f"nhận được: '{line_X}'"
                    )

        elif shape == "Mặt phẳng":
            coeffs = [
                data.get('plane_a', '').strip(),
                data.get('plane_b', '').strip(),
                data.get('plane_c', '').strip(),
                data.get('plane_d', '').strip()
            ]

            valid_coeffs = [c for c in coeffs if c]
            if len(valid_coeffs) == 0:
                issues.append(f"❌ {group_name}: Tất cả hệ số mặt phẳng đều trống")

        elif shape == "Đường tròn":
            center = data.get('circle_center', '').strip()
            radius = data.get('circle_radius', '').strip()

            if not center:
                issues.append(f"❌ {group_name}: Tâm đường tròn trống")
            else:
                coords = center.split(',')
                if len(coords) != 2:
                    issues.append(
                        f"❌ {group_name}: Tâm đường tròn cần 2 tọa độ, "
                        f"nhận được: '{center}'"
                    )

            if not radius:
                issues.append(f"❌ {group_name}: Bán kính đường tròn trống")
            else:
                try:
                    r_value = float(radius)
                    if r_value <= 0:
                        issues.append(f"❌ {group_name}: Bán kính phải > 0")
                except:
                    pass

        elif shape == "Mặt cầu":
            center = data.get('sphere_center', '').strip()
            radius = data.get('sphere_radius', '').strip()

            if not center:
                issues.append(f"❌ {group_name}: Tâm mặt cầu trống")
            else:
                coords = center.split(',')
                if len(coords) != 3:
                    issues.append(
                        f"❌ {group_name}: Tâm mặt cầu cần 3 tọa độ, "
                        f"nhận được: '{center}'"
                    )

            if not radius:
                issues.append(f"❌ {group_name}: Bán kính mặt cầu trống")
            else:
                try:
                    r_value = float(radius)
                    if r_value <= 0:
                        issues.append(f"❌ {group_name}: Bán kính phải > 0")
                except:
                    pass

        return issues

    # ========== END VALIDATION METHODS ==========

    def read_excel_streaming_single_workbook(self, file_path: str, chunksize: int = None) -> Iterator[pd.DataFrame]:
        if chunksize is None:
            chunksize = self.estimate_optimal_chunksize(file_path)
        try:
            import openpyxl
            print(f"🚀 PHƯƠNG ÁN A: Single-workbook streaming")
            print(f"📁 File: {os.path.basename(file_path)}")
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            max_row = ws.max_row if hasattr(ws, 'max_row') and ws.max_row else 0
            max_col = ws.max_column if hasattr(ws, 'max_column') else 0
            total_rows = max(0, max_row - 1)
            self._enforce_row_limit(total_rows)
            print(f"📊 Dimensions: {total_rows:,} rows × {max_col} columns")
            print(f"⚡ Chunk size: {chunksize:,} rows")
            print(f"📦 Estimated chunks: {(total_rows + chunksize - 1) // chunksize}")
            try:
                header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
                columns = [str(cell) if cell is not None else f"Col_{i}" for i, cell in enumerate(header_row)]
            except Exception:
                columns = [f"Col_{i}" for i in range(max_col)]
            current_row = 2
            chunk_count = 0
            while current_row <= max_row and not self.processing_cancelled:
                try:
                    end_row = min(current_row + chunksize - 1, max_row)
                    print(f"⚡ Reading chunk {chunk_count + 1}: rows {current_row:,}-{end_row:,}")
                    chunk_data = []
                    for row in ws.iter_rows(min_row=current_row, max_row=end_row, values_only=True):
                        row_data = [str(cell) if cell is not None else "" for cell in row]
                        chunk_data.append(row_data)
                    if chunk_data:
                        chunk_df = pd.DataFrame(chunk_data, columns=columns).fillna('')
                        yield chunk_df
                        del chunk_df, chunk_data
                        if chunk_count % 10 == 0 and chunk_count > 0:
                            gc.collect()
                            print(f"🧹 Cleanup checkpoint: Memory {self.get_memory_usage():.1f}MB")
                    current_row = end_row + 1
                    chunk_count += 1
                except Exception as e:
                    print(f"❌ Error reading chunk {chunk_count + 1}: {e}")
                    current_row += chunksize
                    continue
            wb.close()
            print(f"✅ Single-workbook streaming completed: {chunk_count} chunks")
        except Exception as e:
            raise Exception(f"Lỗi streaming single-workbook: {str(e)}")

    def process_large_excel_fast(self, file_path: str, shape_a: str, shape_b: str,
                                 operation: str, dimension_a: str, dimension_b: str,
                                 output_path: str, progress_callback: Callable = None) -> Tuple[int, int, str]:
        self.processing_cancelled = False
        success_count = 0
        error_count = 0
        processed_count = 0
        start_time = time.time()
        temp_results_file = f"{output_path}.temp_results"
        from services.geometry.geometry_service import GeometryService
        try:
            print(f"🚀 PHƯƠNG ÁN A - HIGH-SPEED processing: {os.path.basename(file_path)}")
            has_keylog, keylog_col_name, keylog_col_index = self._detect_keylog_column_strict(file_path)
            total_rows = self._get_actual_total_rows(file_path)
            self._enforce_row_limit(total_rows)
            service = GeometryService(self.config)
            service.set_current_shapes(shape_a, shape_b)
            service.set_kich_thuoc(dimension_a, dimension_b)
            service.set_current_operation(operation)
            chunk_size = self.estimate_optimal_chunksize(file_path)
            print(f"⚡ Optimized chunk size: {chunk_size:,} rows")
            print(f"🎯 Target: {total_rows:,} rows at 400+ rows/sec")
            results_buffer = []
            buffer_size = 5000
            chunk_count = 0
            last_speed_check = time.time()
            for chunk_df in self.read_excel_streaming_single_workbook(file_path, chunk_size):
                if self.processing_cancelled:
                    break
                chunk_count += 1
                chunk_start = time.time()
                chunk_results = []
                for index, row in chunk_df.iterrows():
                    try:
                        if self.processing_cancelled:
                            break
                        data_a = self._extract_shape_data_fast(row, shape_a, 'A')
                        data_b = self._extract_shape_data_fast(row, shape_b, 'B') if shape_b else {}
                        if processed_count % 10000 == 0 and processed_count > 0:
                            service = GeometryService(self.config)
                            service.set_current_shapes(shape_a, shape_b)
                            service.set_kich_thuoc(dimension_a, dimension_b)
                            service.set_current_operation(operation)
                        service.thuc_thi_tat_ca(data_a, data_b)
                        result = service.generate_final_result()
                        chunk_results.append(result)
                        success_count += 1
                    except Exception as e:
                        chunk_results.append(f"LỖI: {str(e)}")
                        error_count += 1
                    processed_count += 1
                results_buffer.extend(chunk_results)
                if len(results_buffer) >= buffer_size:
                    self._write_results_buffer_fast(temp_results_file, results_buffer)
                    results_buffer = []
                chunk_time = time.time() - chunk_start
                chunk_speed = len(chunk_df) / chunk_time if chunk_time >= 0.5 else None
                current_time = time.time()
                elapsed = current_time - start_time
                avg_speed = processed_count / elapsed if elapsed > 0 else 0
                if progress_callback and processed_count % 100 == 0:
                    processed_display = min(processed_count, total_rows)
                    progress_percent = (processed_display / total_rows) * 100 if total_rows > 0 else 0
                    progress_callback(progress_percent, processed_display, total_rows, error_count)
                if current_time - last_speed_check >= 5:
                    processed_display = min(processed_count, total_rows)
                    progress_percent = (processed_display / total_rows) * 100 if total_rows > 0 else 0
                    remaining_rows = total_rows - processed_display
                    eta_seconds = remaining_rows / max(avg_speed, 1e-6) if remaining_rows > 0 else 0
                    eta_minutes = int(eta_seconds // 60)
                    eta_secs = int(eta_seconds % 60)
                    eta_str = f"{eta_minutes:02d}:{eta_secs:02d}"
                    if chunk_speed is not None:
                        speed_display = f"🔥 Speed: {avg_speed:.0f} rows/sec (avg) | {chunk_speed:.0f} rows/sec (current)"
                    else:
                        speed_display = f"🔥 Speed: {avg_speed:.0f} rows/sec (avg)"
                    print(
                        f"{speed_display} | Progress: {processed_display:,}/{total_rows:,} ({progress_percent:.1f}%) | ETA: {eta_str}")
                    last_speed_check = current_time
                if chunk_count % 5 == 0 and self.check_memory_limit():
                    print(f"⚠️ Memory: {self.get_memory_usage():.1f}MB - Quick cleanup")
                    gc.collect()
                del chunk_df
            if results_buffer:
                self._write_results_buffer_fast(temp_results_file, results_buffer)
            print("🔧 Creating final Excel file with strict keylog + Flexio font...")
            final_output = self._create_excel_with_smart_keylog(file_path, temp_results_file, output_path,
                                                                has_keylog, keylog_col_name, keylog_col_index)
            total_time = time.time() - start_time
            final_speed = processed_count / total_time if total_time > 0 else 0
            print(f"🏁 PHƯƠNG ÁN A COMPLETED!")
            print(f"⚡ Final speed: {final_speed:.0f} rows/sec (Target: 400+ rows/sec)")
            print(f"📊 Total: {processed_count:,} rows in {total_time:.1f}s")
            print(f"✅ Success: {success_count:,} | ❌ Errors: {error_count:,}")
            return success_count, error_count, final_output
        except Exception as e:
            raise Exception(f"Lỗi xử lý PHƯƠNG ÁN A: {str(e)}")
        finally:
            try:
                if os.path.exists(temp_results_file):
                    os.remove(temp_results_file)
                    print(f"🧹 Cleaned up temp file: {os.path.basename(temp_results_file)}")
            except Exception as cleanup_err:
                print(f"⚠️ Could not remove temp file: {cleanup_err}")

    def _create_excel_with_smart_keylog(self, original_file: str, temp_results_file: str,
                                        output_path: str, has_keylog: bool, keylog_col_name: str,
                                        keylog_col_index: int) -> str:
        try:
            from openpyxl.styles import Font
            print(f"⚡ SMART KEYLOG Excel creation (strict 'keylog' + Flexio font)...")
            all_results = self._read_temp_results_fast(temp_results_file)
            keylog_col_name = 'keylog'
            if has_keylog:
                print(
                    f"📝 Using existing 'keylog' column (position {keylog_col_index + 1 if keylog_col_index >= 0 else '?'} )")
            else:
                print(f"📝 Creating new 'keylog' column")
            try:
                original_df = pd.read_excel(original_file, dtype=str, keep_default_na=False, engine='openpyxl')
                results_to_add = all_results[:len(original_df)]
                if len(results_to_add) < len(original_df):
                    results_to_add.extend([''] * (len(original_df) - len(results_to_add)))
                if 'keylog' in [c.strip().lower() for c in original_df.columns]:
                    exact_name = next((c for c in original_df.columns if c.strip().lower() == 'keylog'), 'keylog')
                    original_df[exact_name] = results_to_add
                    applied_name = exact_name
                else:
                    original_df['keylog'] = results_to_add
                    applied_name = 'keylog'
                with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                    original_df.to_excel(writer, sheet_name='Results', index=False)
                    worksheet = writer.sheets['Results']
                    header_cells = next(worksheet.iter_rows(min_row=1, max_row=1))
                    keylog_col_idx = None
                    for idx, cell in enumerate(header_cells, start=1):
                        if str(cell.value).strip().lower() == 'keylog':
                            keylog_col_idx = idx
                            break
                    if keylog_col_idx:
                        keylog_font = Font(name="Flexio Fx799VN", size=11, bold=True, color="000000")
                        max_row = worksheet.max_row
                        for row in range(1, max_row + 1):
                            worksheet.cell(row=row, column=keylog_col_idx).font = keylog_font
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if cell.value and len(str(cell.value)) > max_length:
                                    max_length = min(len(str(cell.value)), 40)
                            except:
                                pass
                        worksheet.column_dimensions[column_letter].width = max(max_length + 2, 10)
                print(f"✅ SMART KEYLOG Excel creation completed!")
                return output_path
            except Exception as pandas_error:
                print(f"⚠️ Pandas method failed: {pandas_error}")
                return self._create_excel_openpyxl_smart_keylog(original_file, all_results, output_path,
                                                                has_keylog, 'keylog', keylog_col_index)
        except Exception as e:
            raise Exception(f"Lỗi tạo Excel SMART KEYLOG: {str(e)}")

    def _create_excel_openpyxl_smart_keylog(self, original_file: str, results: List[str], output_path: str,
                                            has_keylog: bool, keylog_col_name: str, keylog_col_index: int) -> str:
        try:
            import openpyxl
            from openpyxl.styles import Font
            print("🔄 Using openpyxl smart keylog method (strict 'keylog')...")
            source_wb = openpyxl.load_workbook(original_file, read_only=True, data_only=True)
            source_ws = source_wb.active
            output_wb = openpyxl.Workbook()
            output_ws = output_wb.active
            output_ws.title = "Results"
            header_row = next(source_ws.iter_rows(min_row=1, max_row=1, values_only=True))
            header_list = list(header_row)
            target_col_index = None
            for i, cell in enumerate(header_list):
                if cell is not None and str(cell).strip().lower() == 'keylog':
                    target_col_index = i
                    break
            if target_col_index is None:
                final_header = header_list + ['keylog']
                target_col_index = len(header_list)
                print(f"📝 Will create new column 'keylog' at position {target_col_index + 1}")
            else:
                final_header = header_list
                print(f"📝 Will update existing column 'keylog' at position {target_col_index + 1}")
            for col_idx, header_cell in enumerate(final_header):
                cell_value = str(header_cell) if header_cell is not None else ""
                output_ws.cell(row=1, column=col_idx + 1, value=cell_value)
            row_count = 2
            result_idx = 0
            for data_row in source_ws.iter_rows(min_row=2, values_only=True):
                if result_idx >= len(results):
                    break
                data_list = list(data_row)
                if len(data_list) < len(final_header):
                    data_list += [""] * (len(final_header) - len(data_list))
                data_list[target_col_index] = results[result_idx]
                for col_idx, cell_value in enumerate(data_list):
                    value = str(cell_value) if cell_value is not None else ""
                    output_ws.cell(row=row_count, column=col_idx + 1, value=value)
                row_count += 1
                result_idx += 1
                if row_count % 5000 == 0:
                    gc.collect()
            keylog_font = Font(name="Flexio Fx799VN", size=11, bold=True, color="000000")
            col = target_col_index + 1
            max_row = output_ws.max_row
            for row in range(1, max_row + 1):
                output_ws.cell(row=row, column=col).font = keylog_font
            source_wb.close()
            output_wb.save(output_path)
            output_wb.close()
            print(f"✅ Openpyxl smart keylog method completed!")
            return output_path
        except Exception as e:
            raise Exception(f"Lỗi openpyxl smart keylog fallback: {str(e)}")

    def _extract_shape_data_fast(self, row: pd.Series, shape_type: str, group: str) -> Dict:
        data_dict = {}
        if not shape_type:
            return data_dict
        if group == 'A':
            if shape_type == "Điểm":
                data_dict['point_input'] = str(row.get('data_A', '')).strip()
            elif shape_type == "Đường thẳng":
                data_dict['line_A1'] = str(row.get('d_P_data_A', '')).strip()
                data_dict['line_X1'] = str(row.get('d_V_data_A', '')).strip()
            elif shape_type == "Mặt phẳng":
                data_dict['plane_a'] = str(row.get('P1_a', '')).strip()
                data_dict['plane_b'] = str(row.get('P1_b', '')).strip()
                data_dict['plane_c'] = str(row.get('P1_c', '')).strip()
                data_dict['plane_d'] = str(row.get('P1_d', '')).strip()
            elif shape_type == "Đường tròn":
                data_dict['circle_center'] = str(row.get('C_data_I1', '')).strip()
                data_dict['circle_radius'] = str(row.get('C_data_R1', '')).strip()
            elif shape_type == "Mặt cầu":
                data_dict['sphere_center'] = str(row.get('S_data_I1', '')).strip()
                data_dict['sphere_radius'] = str(row.get('S_data_R1', '')).strip()
        else:
            if shape_type == "Điểm":
                data_dict['point_input'] = str(row.get('data_B', '')).strip()
            elif shape_type == "Đường thẳng":
                data_dict['line_A2'] = str(row.get('d_P_data_B', '')).strip()
                data_dict['line_X2'] = str(row.get('d_V_data_B', '')).strip()
            elif shape_type == "Mặt phẳng":
                data_dict['plane_a'] = str(row.get('P2_a', '')).strip()
                data_dict['plane_b'] = str(row.get('P2_b', '')).strip()
                data_dict['plane_c'] = str(row.get('P2_c', '')).strip()
                data_dict['plane_d'] = str(row.get('P2_d', '')).strip()
            elif shape_type == "Đường tròn":
                data_dict['circle_center'] = str(row.get('C_data_I2', '')).strip()
                data_dict['circle_radius'] = str(row.get('C_data_R2', '')).strip()
            elif shape_type == "Mặt cầu":
                data_dict['sphere_center'] = str(row.get('S_data_I2', '')).strip()
                data_dict['sphere_radius'] = str(row.get('S_data_R2', '')).strip()
        return data_dict

    def _write_results_buffer_fast(self, temp_file: str, results: List[str]):
        try:
            mode = 'a' if os.path.exists(temp_file) else 'w'
            with open(temp_file, mode, encoding='utf-8', buffering=16384) as f:
                f.write('\n'.join(results) + '\n')
        except Exception as e:
            print(f"⚠️ Warning: Fast buffer write failed: {e}")

    def _read_temp_results_fast(self, temp_file: str) -> List[str]:
        try:
            with open(temp_file, 'r', encoding='utf-8', buffering=16384) as f:
                return f.read().strip().split('\n')
        except Exception:
            return []

    def _create_excel_direct_fast(self, original_file: str, temp_results_file: str, output_path: str) -> str:
        return self._create_excel_with_smart_keylog(original_file, temp_results_file, output_path,
                                                    *self._detect_keylog_column_strict(original_file))

    def _create_excel_openpyxl_fast(self, original_file: str, results: List[str], output_path: str) -> str:
        has_keylog, keylog_col_name, keylog_col_index = self._detect_keylog_column_strict(original_file)
        return self._create_excel_openpyxl_smart_keylog(original_file, results, output_path,
                                                        has_keylog, keylog_col_name, keylog_col_index)

    def process_large_excel_safe(self, file_path: str, shape_a: str, shape_b: str,
                                 operation: str, dimension_a: str, dimension_b: str,
                                 output_path: str, progress_callback: Callable = None) -> Tuple[int, int, str]:
        """
        MAIN ENTRY POINT với PRE-VALIDATION
        """

        print(f"🔥 Switching to LARGE FILE MODE for: {os.path.basename(file_path)}")

        # ===== VALIDATION TRƯỚC KHI CHUNKING =====
        validation_result = self._validate_first_row_before_chunking(
            file_path, shape_a, shape_b
        )

        if not validation_result['valid']:
            # Tạo error message chi tiết
            error_msg = "❌ Validation failed! Cannot process file.\n\n"
            error_msg += "Issues found:\n"
            for issue in validation_result['issues']:
                error_msg += f"  • {issue}\n"

            if validation_result['warnings']:
                error_msg += "\nWarnings:\n"
                for warning in validation_result['warnings']:
                    error_msg += f"  • {warning}\n"

            if validation_result['sample_data']:
                error_msg += "\n📊 Sample data from row 1:\n"
                for group, data in validation_result['sample_data'].items():
                    error_msg += f"\n{group}:\n"
                    for key, val in data.items():
                        error_msg += f"  {key}: '{val}'\n"

            raise ValueError(error_msg)

        print("✅ Validation passed - Starting chunked processing...")

        # ===== TIẾP TỤC XỬ LÝ =====
        return self.process_large_excel_fast(file_path, shape_a, shape_b, operation,
                                             dimension_a, dimension_b, output_path, progress_callback)

    def validate_large_file_structure(self, file_path: str, shape_a: str, shape_b: str = None) -> Dict[str, Any]:
        try:
            actual_rows = self._get_actual_total_rows(file_path)
            file_size_mb = os.path.getsize(file_path) / (1024 * 1024)
            header_df = pd.read_excel(file_path, nrows=0, engine='openpyxl')
            columns = header_df.columns.tolist()
            has_keylog, keylog_col_name, keylog_col_index = self._detect_keylog_column_strict(file_path)
            over_limit = actual_rows > self.max_rows_allowed
            required_columns_A = self._get_required_columns(shape_a, 'A')
            required_columns_B = self._get_required_columns(shape_b, 'B') if shape_b else []
            missing_columns = [col for col in (required_columns_A + required_columns_B) if col not in columns]
            return {
                'valid': len(missing_columns) == 0 and not over_limit,
                'file_size_mb': file_size_mb,
                'actual_rows': actual_rows,
                'columns': columns,
                'missing_columns': missing_columns,
                'has_keylog_column': has_keylog,
                'keylog_column_name': keylog_col_name,
                'keylog_column_index': keylog_col_index,
                'recommended_chunk_size': self.estimate_optimal_chunksize(file_path),
                'max_rows_allowed': self.max_rows_allowed,
                'over_row_limit': over_limit,
                'validation_method': 'phương_án_A_strict_keylog_with_prevalidation',
                'warning': f'File vượt quá giới hạn {self.max_rows_allowed:,} dòng' if over_limit else ''
            }
        except Exception as e:
            return {'valid': False, 'error': f'Lỗi kiểm tra file: {str(e)}'}

    def _get_required_columns(self, shape: str, group: str) -> List[str]:
        if not shape:
            return []
        if group == 'A':
            mapping = {
                "Điểm": ["data_A"],
                "Đường thẳng": ["d_P_data_A", "d_V_data_A"],
                "Mặt phẳng": ["P1_a", "P1_b", "P1_c", "P1_d"],
                "Đường tròn": ["C_data_I1", "C_data_R1"],
                "Mặt cầu": ["S_data_I1", "S_data_R1"]
            }
        else:
            mapping = {
                "Điểm": ["data_B"],
                "Đường thẳng": ["d_P_data_B", "d_V_data_B"],
                "Mặt phẳng": ["P2_a", "P2_b", "P2_c", "P2_d"],
                "Đường tròn": ["C_data_I2", "C_data_R2"],
                "Mặt cầu": ["S_data_I2", "S_data_R2"]
            }
        return mapping.get(shape, [])

    def get_processing_statistics(self) -> Dict[str, Any]:
        return {
            'memory_usage_mb': self.get_memory_usage(),
            'memory_limit_mb': self.max_memory_mb,
            'memory_usage_percent': (self.get_memory_usage() / self.max_memory_mb) * 100,
            'emergency_cleanup_triggered': self.emergency_cleanup,
            'processing_cancelled': self.processing_cancelled,
            'recommended_max_chunksize': 5000 if self.get_memory_usage() > 800 else 7000,
            'max_rows_allowed': self.max_rows_allowed,
            'processing_method': 'phương_án_A_with_prevalidation',
            'target_speed_rows_per_sec': 400,
            'optimizations': ['single_workbook_open', 'strict_keylog', 'flexio_font',
                              'large_chunks', 'minimal_gc', 'batch_io', 'pre_validation']
        }