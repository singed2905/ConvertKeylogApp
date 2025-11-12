import pandas as pd
import json
import os
import openpyxl
from typing import Dict, List, Tuple, Any, Optional
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
import re
from datetime import datetime
from .large_file_processor import LargeFileProcessor


class ExcelProcessor:
    """Excel Processor for ConvertKeylogApp - Enhanced with large file support and first row validation"""

    def __init__(self, config: Dict = None):
        self.config = config or {}
        self.mapping = self._load_mapping()
        self.large_file_processor = LargeFileProcessor(config)
        self.large_file_threshold_mb = 20
        self.large_file_threshold_rows = 50000

    def _load_mapping(self) -> Dict:
        """Load Excel mapping configuration from config or fallback"""
        try:
            if self.config and 'geometry' in self.config:
                geometry_config = self.config['geometry']
                if 'excel_mapping' in geometry_config:
                    return geometry_config['excel_mapping']

            mapping_file = "config/geometry_mode/geometry_excel_mapping.json"
            if os.path.exists(mapping_file):
                with open(mapping_file, 'r', encoding='utf-8') as f:
                    return json.load(f)

        except Exception as e:
            print(f"Warning: Could not load Excel mapping: {e}")

        return self._get_default_mapping()

    def _get_default_mapping(self) -> Dict:
        """Default Excel mapping matching TL structure"""
        return {
            "group_a_mapping": {
                "Điểm": {
                    "required_columns": ["data_A"],
                    "columns": {
                        "point_input": {"excel_column": "data_A"}
                    }
                },
                "Đường thẳng": {
                    "required_columns": ["d_P_data_A", "d_V_data_A"],
                    "columns": {
                        "line_A1": {"excel_column": "d_P_data_A"},
                        "line_X1": {"excel_column": "d_V_data_A"}
                    }
                },
                "Mặt phẳng": {
                    "required_columns": ["P1_a", "P1_b", "P1_c", "P1_d"],
                    "columns": {
                        "plane_a": {"excel_column": "P1_a"},
                        "plane_b": {"excel_column": "P1_b"},
                        "plane_c": {"excel_column": "P1_c"},
                        "plane_d": {"excel_column": "P1_d"}
                    }
                },
                "Đường tròn": {
                    "required_columns": ["C_data_I1", "C_data_R1"],
                    "columns": {
                        "circle_center": {"excel_column": "C_data_I1"},
                        "circle_radius": {"excel_column": "C_data_R1"}
                    }
                },
                "Mặt cầu": {
                    "required_columns": ["S_data_I1", "S_data_R1"],
                    "columns": {
                        "sphere_center": {"excel_column": "S_data_I1"},
                        "sphere_radius": {"excel_column": "S_data_R1"}
                    }
                }
            },
            "group_b_mapping": {
                "Điểm": {
                    "required_columns": ["data_B"],
                    "columns": {
                        "point_input": {"excel_column": "data_B"}
                    }
                },
                "Đường thẳng": {
                    "required_columns": ["d_P_data_B", "d_V_data_B"],
                    "columns": {
                        "line_A2": {"excel_column": "d_P_data_B"},
                        "line_X2": {"excel_column": "d_V_data_B"}
                    }
                },
                "Mặt phẳng": {
                    "required_columns": ["P2_a", "P2_b", "P2_c", "P2_d"],
                    "columns": {
                        "plane_a": {"excel_column": "P2_a"},
                        "plane_b": {"excel_column": "P2_b"},
                        "plane_c": {"excel_column": "P2_c"},
                        "plane_d": {"excel_column": "P2_d"}
                    }
                },
                "Đường tròn": {
                    "required_columns": ["C_data_I2", "C_data_R2"],
                    "columns": {
                        "circle_center": {"excel_column": "C_data_I2"},
                        "circle_radius": {"excel_column": "C_data_R2"}
                    }
                },
                "Mặt cầu": {
                    "required_columns": ["S_data_I2", "S_data_R2"],
                    "columns": {
                        "sphere_center": {"excel_column": "S_data_I2"},
                        "sphere_radius": {"excel_column": "S_data_R2"}
                    }
                }
            }
        }

    def is_large_file(self, file_path: str) -> Tuple[bool, Dict[str, Any]]:
        """Check if file is too large for normal processing"""
        try:
            file_size_mb = os.path.getsize(file_path) / (1024 * 1024)

            wb = openpyxl.load_workbook(file_path, read_only=True)
            ws = wb.active
            estimated_rows = ws.max_row - 1 if hasattr(ws, 'max_row') else 0
            wb.close()

            is_large = (file_size_mb > self.large_file_threshold_mb or
                        estimated_rows > self.large_file_threshold_rows)

            return is_large, {
                'file_size_mb': file_size_mb,
                'estimated_rows': estimated_rows,
                'recommended_processor': 'large_file' if is_large else 'normal',
                'recommended_chunk_size': self.large_file_processor.estimate_optimal_chunksize(file_path)
            }

        except Exception as e:
            return False, {'error': f'Không thể phân tích file: {str(e)}'}

    def validate_first_row_data(self, file_path: str, shape_a: str,
                                shape_b: str = None) -> Dict[str, Any]:
        """
        Validate dòng dữ liệu đầu tiên trước khi batch processing

        Returns:
            {
                'valid': bool,
                'has_structure': bool,
                'has_data': bool,
                'issues': List[str],
                'warnings': List[str],
                'sample_data': Dict
            }
        """
        result = {
            'valid': True,
            'has_structure': False,
            'has_data': False,
            'issues': [],
            'warnings': [],
            'sample_data': {}
        }

        try:
            # Kiểm tra file có quá lớn không
            is_large, file_info = self.is_large_file(file_path)

            if is_large:
                return self._validate_first_row_large_file(
                    file_path, shape_a, shape_b, file_info
                )

            # Đọc file bình thường
            df = pd.read_excel(file_path)
            df.columns = df.columns.str.strip()

            # Validate cấu trúc (header)
            is_valid_structure, missing_cols = self.validate_excel_structure(
                df, shape_a, shape_b
            )

            result['has_structure'] = is_valid_structure

            if not is_valid_structure:
                result['valid'] = False
                result['issues'].append(
                    f"❌ Thiếu các cột bắt buộc: {', '.join(missing_cols)}"
                )
                return result

            # Kiểm tra có ít nhất 1 dòng data không
            if len(df) == 0:
                result['valid'] = False
                result['issues'].append("❌ File Excel không có dữ liệu (chỉ có header)")
                return result

            # Validate dòng đầu tiên
            first_row = df.iloc[0]

            # Extract data từ dòng 1
            data_a = self.extract_shape_data(first_row, shape_a, 'A')
            data_b = self.extract_shape_data(first_row, shape_b, 'B') if shape_b else {}

            # Validate data Group A
            has_data_a = any(str(v).strip() for v in data_a.values())
            issues_a = self._validate_shape_data_detailed(data_a, shape_a, "Nhóm A")

            if not has_data_a:
                result['issues'].append(
                    f"❌ Dòng 1 - Nhóm A ({shape_a}): Không có dữ liệu"
                )
            elif issues_a:
                result['issues'].extend(issues_a)

            # Validate data Group B (nếu có)
            if shape_b:
                has_data_b = any(str(v).strip() for v in data_b.values())
                issues_b = self._validate_shape_data_detailed(data_b, shape_b, "Nhóm B")

                if not has_data_b:
                    result['issues'].append(
                        f"❌ Dòng 1 - Nhóm B ({shape_b}): Không có dữ liệu"
                    )
                elif issues_b:
                    result['issues'].extend(issues_b)

            # Tổng hợp kết quả
            result['has_data'] = has_data_a and (not shape_b or has_data_b)
            result['valid'] = result['has_structure'] and result['has_data'] and len(result['issues']) == 0

            # Thêm sample data để preview
            result['sample_data'] = {
                'group_a': data_a,
                'group_b': data_b if shape_b else None
            }

            # Thêm warnings
            if len(df) < 5:
                result['warnings'].append(
                    f"⚠️ File chỉ có {len(df)} dòng dữ liệu"
                )

            return result

        except Exception as e:
            result['valid'] = False
            result['issues'].append(f"❌ Lỗi đọc file: {str(e)}")
            return result

    def _validate_first_row_large_file(self, file_path: str, shape_a: str,
                                       shape_b: str, file_info: Dict) -> Dict:
        """Validate dòng đầu cho file lớn (dùng openpyxl)"""
        result = {
            'valid': True,
            'has_structure': False,
            'has_data': False,
            'issues': [],
            'warnings': [],
            'sample_data': {}
        }

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active

            # Header
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            columns = [str(cell).strip() if cell else "" for cell in header_row]

            # Dòng 1
            first_data_row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True), None)

            wb.close()

            if not first_data_row:
                result['valid'] = False
                result['issues'].append("❌ File không có dòng dữ liệu")
                return result

            # Tạo pandas Series từ dòng đầu
            first_row = pd.Series(first_data_row, index=columns)

            # Check structure
            is_valid_structure, missing_cols = self._check_columns_exist(
                columns, shape_a, shape_b
            )

            result['has_structure'] = is_valid_structure

            if not is_valid_structure:
                result['valid'] = False
                result['issues'].append(
                    f"❌ Thiếu cột: {', '.join(missing_cols)}"
                )
                return result

            # Validate data
            data_a = self.extract_shape_data(first_row, shape_a, 'A')
            data_b = self.extract_shape_data(first_row, shape_b, 'B') if shape_b else {}

            has_data_a = any(str(v).strip() for v in data_a.values())
            issues_a = self._validate_shape_data_detailed(data_a, shape_a, "Nhóm A")

            if not has_data_a or issues_a:
                result['issues'].extend(issues_a or [f"❌ Nhóm A: Không có dữ liệu"])

            if shape_b:
                has_data_b = any(str(v).strip() for v in data_b.values())
                issues_b = self._validate_shape_data_detailed(data_b, shape_b, "Nhóm B")

                if not has_data_b or issues_b:
                    result['issues'].extend(issues_b or [f"❌ Nhóm B: Không có dữ liệu"])

            result['has_data'] = has_data_a and (not shape_b or has_data_b)
            result['valid'] = result['has_structure'] and result['has_data'] and len(result['issues']) == 0

            result['sample_data'] = {
                'group_a': data_a,
                'group_b': data_b if shape_b else None
            }

            result['warnings'].append(
                f"⚠️ File lớn ({file_info['file_size_mb']:.1f}MB, "
                f"~{file_info['estimated_rows']:,} dòng) - "
                f"Sẽ dùng chế độ xử lý file lớn"
            )

            return result

        except Exception as e:
            result['valid'] = False
            result['issues'].append(f"❌ Lỗi: {str(e)}")
            return result

    def _check_columns_exist(self, columns: List[str], shape_a: str,
                             shape_b: str = None) -> Tuple[bool, List[str]]:
        """Kiểm tra các cột có tồn tại không"""
        missing_columns = []

        if shape_a in self.mapping['group_a_mapping']:
            required_cols = self.mapping['group_a_mapping'][shape_a]['required_columns']
            for col in required_cols:
                if col not in columns:
                    missing_columns.append(f"Nhóm A - {col}")

        if shape_b and shape_b in self.mapping['group_b_mapping']:
            required_cols = self.mapping['group_b_mapping'][shape_b]['required_columns']
            for col in required_cols:
                if col not in columns:
                    missing_columns.append(f"Nhóm B - {col}")

        return len(missing_columns) == 0, missing_columns

    def _validate_shape_data_detailed(self, data: Dict, shape: str,
                                      group_name: str) -> List[str]:
        """Validate chi tiết data của 1 shape"""
        issues = []

        if shape == "Điểm":
            point_input = data.get('point_input', '').strip()

            if not point_input:
                issues.append(f"❌ {group_name}: Tọa độ điểm trống")
            else:
                coords = point_input.split(',')
                if len(coords) < 2:
                    issues.append(
                        f"❌ {group_name}: Điểm cần ít nhất 2 tọa độ, "
                        f"nhận được: '{point_input}'"
                    )
                else:
                    for i, coord in enumerate(coords):
                        if not self._is_valid_number_or_expression(coord.strip()):
                            issues.append(
                                f"❌ {group_name}: Tọa độ {i + 1} không hợp lệ: '{coord}'"
                            )
                            break

        elif shape == "Đường thẳng":
            line_A = data.get('line_A1') or data.get('line_A2', '')
            line_X = data.get('line_X1') or data.get('line_X2', '')

            if not line_A.strip():
                issues.append(f"❌ {group_name}: Điểm trên đường thẳng trống")
            else:
                coords_A = line_A.split(',')
                if len(coords_A) != 3:
                    issues.append(
                        f"❌ {group_name}: Điểm đường thẳng cần 3 tọa độ, "
                        f"nhận được: '{line_A}'"
                    )

            if not line_X.strip():
                issues.append(f"❌ {group_name}: Vector phương trống")
            else:
                coords_X = line_X.split(',')
                if len(coords_X) != 3:
                    issues.append(
                        f"❌ {group_name}: Vector phương cần 3 tọa độ, "
                        f"nhận được: '{line_X}'"
                    )

        elif shape == "Mặt phẳng":
            coeffs = [
                data.get('plane_a', '').strip(),
                data.get('plane_b', '').strip(),
                data.get('plane_c', '').strip(),
                data.get('plane_d', '').strip()
            ]

            valid_coeffs = [c for c in coeffs if c]

            if len(valid_coeffs) == 0:
                issues.append(f"❌ {group_name}: Tất cả hệ số mặt phẳng đều trống")
            elif len(valid_coeffs) < 4:
                issues.append(
                    f"⚠️ {group_name}: Chỉ có {len(valid_coeffs)}/4 hệ số "
                    f"(a, b, c, d). Có thể gây lỗi tính toán."
                )

            coeff_names = ['a', 'b', 'c', 'd']
            for i, coeff in enumerate(coeffs):
                if coeff and not self._is_valid_number_or_expression(coeff):
                    issues.append(
                        f"❌ {group_name}: Hệ số {coeff_names[i]} không hợp lệ: '{coeff}'"
                    )

        elif shape == "Đường tròn":
            center = data.get('circle_center', '').strip()
            radius = data.get('circle_radius', '').strip()

            if not center:
                issues.append(f"❌ {group_name}: Tâm đường tròn trống")
            else:
                coords = center.split(',')
                if len(coords) != 2:
                    issues.append(
                        f"❌ {group_name}: Tâm đường tròn cần 2 tọa độ, "
                        f"nhận được: '{center}'"
                    )

            if not radius:
                issues.append(f"❌ {group_name}: Bán kính đường tròn trống")
            elif not self._is_valid_number_or_expression(radius):
                issues.append(f"❌ {group_name}: Bán kính không hợp lệ: '{radius}'")
            else:
                try:
                    r_value = float(radius)
                    if r_value <= 0:
                        issues.append(f"❌ {group_name}: Bán kính phải > 0, nhận được: {r_value}")
                except:
                    pass

        elif shape == "Mặt cầu":
            center = data.get('sphere_center', '').strip()
            radius = data.get('sphere_radius', '').strip()

            if not center:
                issues.append(f"❌ {group_name}: Tâm mặt cầu trống")
            else:
                coords = center.split(',')
                if len(coords) != 3:
                    issues.append(
                        f"❌ {group_name}: Tâm mặt cầu cần 3 tọa độ, "
                        f"nhận được: '{center}'"
                    )

            if not radius:
                issues.append(f"❌ {group_name}: Bán kính mặt cầu trống")
            elif not self._is_valid_number_or_expression(radius):
                issues.append(f"❌ {group_name}: Bán kính không hợp lệ: '{radius}'")
            else:
                try:
                    r_value = float(radius)
                    if r_value <= 0:
                        issues.append(f"❌ {group_name}: Bán kính phải > 0, nhận được: {r_value}")
                except:
                    pass

        return issues

    def _is_valid_number_or_expression(self, value: str) -> bool:
        """Kiểm tra value có phải số hoặc biểu thức toán hợp lệ"""
        if not value or not value.strip():
            return False

        value = value.strip()

        try:
            float(value)
            return True
        except:
            pass

        math_chars = set('0123456789+-*/().sqrtsincotan π')
        clean_value = value.lower().replace(' ', '')

        for char in clean_value:
            if char not in math_chars and not char.isalpha():
                return False

        return True

    def read_excel_data(self, file_path: str) -> pd.DataFrame:
        """Read Excel file and normalize data"""
        try:
            is_large, file_info = self.is_large_file(file_path)

            if is_large:
                raise Exception(
                    f"File quá lớn cho phương thức thông thường!\n"
                    f"Kích thước: {file_info.get('file_size_mb', 0):.1f}MB\n"
                    f"Dòng ước tính: {file_info.get('estimated_rows', 0):,}\n\n"
                    f"Vui lòng sử dụng chế độ xử lý file lớn."
                )

            df = pd.read_excel(file_path)
            df.columns = df.columns.str.strip()
            return df
        except Exception as e:
            raise Exception(f"Không thể đọc file Excel: {str(e)}")

    def validate_excel_structure(self, df: pd.DataFrame, shape_a: str, shape_b: str = None) -> Tuple[bool, List[str]]:
        """Validate Excel structure against selected shapes"""
        missing_columns = []

        if shape_a in self.mapping['group_a_mapping']:
            required_cols = self.mapping['group_a_mapping'][shape_a]['required_columns']
            for col in required_cols:
                if col not in df.columns:
                    missing_columns.append(f"Nhóm A - {col}")

        if shape_b and shape_b in self.mapping['group_b_mapping']:
            required_cols = self.mapping['group_b_mapping'][shape_b]['required_columns']
            for col in required_cols:
                if col not in df.columns:
                    missing_columns.append(f"Nhóm B - {col}")

        return len(missing_columns) == 0, missing_columns

    def validate_large_file_structure(self, file_path: str, shape_a: str, shape_b: str = None) -> Dict[str, Any]:
        """Validate large file structure without loading entire file"""
        return self.large_file_processor.validate_large_file_structure(file_path, shape_a, shape_b)

    def extract_shape_data(self, row: pd.Series, shape_type: str, group: str) -> Dict:
        """Extract data for specific shape from Excel row"""
        if group == 'A':
            shape_mapping = self.mapping['group_a_mapping'].get(shape_type, {})
        else:
            shape_mapping = self.mapping['group_b_mapping'].get(shape_type, {})

        data_dict = {}
        for field, config in shape_mapping.get('columns', {}).items():
            excel_column = config.get('excel_column')
            if excel_column and excel_column in row.index:
                value = row[excel_column]
                if pd.isna(value):
                    data_dict[field] = ""
                else:
                    data_dict[field] = str(value).strip()
            else:
                data_dict[field] = ""

        return data_dict

    def process_large_excel_file(self, file_path: str, shape_a: str, shape_b: str,
                                 operation: str, dimension_a: str, dimension_b: str,
                                 output_path: str, progress_callback: callable = None) -> Tuple[int, int, str]:
        """Process large Excel files using specialized processor"""
        try:
            print(f"🔥 Switching to LARGE FILE MODE for: {os.path.basename(file_path)}")

            return self.large_file_processor.process_large_excel_safe(
                file_path, shape_a, shape_b, operation, dimension_a, dimension_b,
                output_path, progress_callback
            )

        except Exception as e:
            raise Exception(f"Lỗi xử lý file lớn: {str(e)}")

    def export_results(self, original_df: pd.DataFrame, encoded_results: List[str], output_path: str) -> str:
        """Export results with Excel formatting"""
        try:
            result_df = original_df.copy()

            keylog_column = None
            for col in result_df.columns:
                if col.strip().lower() == 'keylog':
                    keylog_column = col
                    break

            if keylog_column:
                result_df[keylog_column] = encoded_results
            else:
                result_df['Kết quả mã hóa'] = encoded_results

            os.makedirs(os.path.dirname(output_path), exist_ok=True)

            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                result_df.to_excel(writer, index=False, sheet_name='Results')
                worksheet = writer.sheets['Results']
                self._format_results_worksheet(worksheet, result_df, keylog_column)

            return output_path

        except Exception as e:
            raise Exception(f"Không thể xuất file kết quả: {str(e)}")

    def _format_results_worksheet(self, worksheet, df, keylog_column=None):
        """Format Excel worksheet with colors and fonts"""
        try:
            header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='2E86AB', end_color='2E86AB', fill_type='solid')
            data_font = Font(name='Arial', size=10)
            result_font = Font(name='Arial', size=10, bold=True, color='2E7D32')

            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill

            keylog_col_name = keylog_column if keylog_column else 'Kết quả mã hóa'
            keylog_col_idx = None
            for idx, col_name in enumerate(df.columns):
                if keylog_col_name in str(col_name):
                    keylog_col_idx = idx
                    break

            max_format_rows = min(len(df) + 2, 10000)
            for row in range(2, max_format_rows):
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row, column=col)

                    if keylog_col_idx is not None and col == keylog_col_idx + 1:
                        cell.font = result_font
                    else:
                        cell.font = data_font

            for col_idx, column in enumerate(worksheet.columns):
                if col_idx > 20:
                    break

                max_length = 0
                column_letter = get_column_letter(column[0].column)

                for i, cell in enumerate(column):
                    if i > 100:
                        break
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

        except Exception as e:
            print(f"Warning: Could not format worksheet: {e}")

    def get_total_rows(self, file_path: str) -> int:
        """Get total number of rows in Excel file"""
        try:
            file_size_mb = os.path.getsize(file_path) / (1024 * 1024)

            if file_size_mb > 10:
                wb = openpyxl.load_workbook(file_path, read_only=True)
                ws = wb.active
                total_rows = ws.max_row - 1 if hasattr(ws, 'max_row') else 0
                wb.close()
                return total_rows
            else:
                df = pd.read_excel(file_path)
                return len(df)
        except Exception:
            return 0

    def get_file_info(self, file_path: str) -> Dict[str, Any]:
        """Get information about Excel file"""
        try:
            is_large, large_file_info = self.is_large_file(file_path)

            if is_large:
                return self._get_large_file_info(file_path, large_file_info)
            else:
                df = self.read_excel_data(file_path)
                file_name = os.path.basename(file_path)

                return {
                    'file_name': file_name,
                    'total_rows': len(df),
                    'total_columns': len(df.columns),
                    'columns': list(df.columns),
                    'file_size': os.path.getsize(file_path),
                    'file_size_mb': os.path.getsize(file_path) / (1024 * 1024),
                    'is_large_file': False,
                    'first_few_rows': df.head(3).to_dict('records') if len(df) > 0 else []
                }
        except Exception as e:
            raise Exception(f"Không thể đọc thông tin file: {str(e)}")

    def _get_large_file_info(self, file_path: str, large_file_info: Dict) -> Dict[str, Any]:
        """Get file info for large files without loading data"""
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            ws = wb.active

            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            columns = [str(cell) for cell in header_row if cell is not None]

            sample_rows = []
            for i, row in enumerate(ws.iter_rows(min_row=2, max_row=5, values_only=True)):
                if i >= 3:
                    break
                row_dict = {col: str(cell) if cell is not None else "" for col, cell in zip(columns, row)}
                sample_rows.append(row_dict)

            wb.close()

            return {
                'file_name': os.path.basename(file_path),
                'total_rows': large_file_info.get('estimated_rows', 0),
                'total_columns': len(columns),
                'columns': columns,
                'file_size': os.path.getsize(file_path),
                'file_size_mb': large_file_info.get('file_size_mb', 0),
                'is_large_file': True,
                'recommended_chunk_size': large_file_info.get('recommended_chunk_size', 500),
                'first_few_rows': sample_rows,
                'warning': 'File lớn - khuyến nghị dùng xử lý chunked'
            }

        except Exception as e:
            raise Exception(f"Không thể phân tích file lớn: {str(e)}")