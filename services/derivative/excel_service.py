import os
import gc
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from typing import List, Tuple, Dict, Optional, Callable, Iterator
from datetime import datetime
import time
import math

try:
    import psutil
    PSUTIL_AVAILABLE = True
except ImportError:
    PSUTIL_AVAILABLE = False
    print("⚠️ Warning: psutil not available, memory monitoring disabled")

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    print("⚠️ Warning: pandas not available, some features disabled")


class ExcelService:


    # ========== CONSTANTS ==========
    MAX_ROWS_ALLOWED = 250_000      # Giới hạn số dòng tối đa
    MAX_MEMORY_MB = 1500            # Giới hạn memory usage
    FILE_SIZE_THRESHOLD_MB = 10     # Ngưỡng file lớn (MB)

    def __init__(self):
        self.file_path = None
        self.file_size_mb = 0
        self.total_rows = 0
        self.processing_cancelled = False
        self.emergency_cleanup = False

    # ========== FILE VALIDATION ==========

    def _validate_file_type(self, file_path: str):
        """Kiểm tra file phải là XLSX"""
        if not file_path.endswith('.xlsx'):
            raise ValueError(
                f"❌ Chỉ hỗ trợ file Excel (.xlsx)\n"
                f"File của bạn: {os.path.basename(file_path)}"
            )

    # ========== MEMORY MONITORING ==========

    def get_memory_usage(self) -> float:
        """Lấy memory usage hiện tại (MB)"""
        if not PSUTIL_AVAILABLE:
            return 0.0
        try:
            process = psutil.Process()
            return process.memory_info().rss / 1024 / 1024
        except:
            return 0.0

    def check_memory_limit(self) -> bool:
        """Kiểm tra có vượt giới hạn memory không"""
        return self.get_memory_usage() > self.MAX_MEMORY_MB

    def emergency_memory_cleanup(self):
        """Emergency cleanup khi memory cao"""
        self.emergency_cleanup = True
        gc.collect()
        if PSUTIL_AVAILABLE:
            print(f"🧹 Emergency cleanup - Memory: {self.get_memory_usage():.1f}MB")

    # ========== CHUNK SIZE OPTIMIZATION ==========

    def estimate_optimal_chunksize(self, file_path: str) -> int:
        """
        Tính chunk size tối ưu dựa trên file size

        Returns:
            int: Recommended chunk size
        """
        try:
            file_size_mb = os.path.getsize(file_path) / (1024 * 1024)

            if file_size_mb > 100:
                return 3000
            elif file_size_mb > 50:
                return 5000
            elif file_size_mb > 20:
                return 7000
            else:
                return 10000
        except Exception:
            return 5000  # Default

    # ========== ROW VALIDATION ==========

    def _get_actual_total_rows(self, file_path: str) -> int:
        """Đếm số dòng thực tế trong file Excel"""
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            actual_total = max(0, (ws.max_row - 1)) if hasattr(ws, 'max_row') and ws.max_row else 0
            wb.close()
            print(f"📊 File có {actual_total:,} dòng dữ liệu")
            return actual_total
        except Exception as e:
            print(f"⚠️ Không đếm được số dòng: {e}")
            return 0

    def _enforce_row_limit(self, total_rows: int):
        """Kiểm tra giới hạn số dòng"""
        if total_rows > self.MAX_ROWS_ALLOWED:
            raise Exception(
                f"❌ File có {total_rows:,} dòng, vượt quá giới hạn {self.MAX_ROWS_ALLOWED:,} dòng.\n"
                f"Vui lòng chia nhỏ file hoặc lọc bớt dữ liệu."
            )

    # ========== PRE-VALIDATION ==========

    def _validate_first_row(self, file_path: str) -> Dict:

        result = {
            'valid': True,
            'issues': [],
            'warnings': [],
            'sample_data': {}
        }

        try:
            print("🔍 Pre-validation: Checking first row...")

            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active

            # Header
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            columns = [str(cell).strip() if cell else "" for cell in header_row]

            # Dòng đầu tiên
            first_data_row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True), None)
            wb.close()

            if not first_data_row:
                result['valid'] = False
                result['issues'].append("❌ File không có dòng dữ liệu")
                return result

            # Kiểm tra cột 'derivative_input'
            deriv_input_idx = None
            for i, col in enumerate(columns):
                if col.lower() == 'derivative_input':
                    deriv_input_idx = i
                    break

            if deriv_input_idx is None:
                result['valid'] = False
                result['issues'].append("❌ Không tìm thấy cột 'derivative_input'")
                return result

            # Kiểm tra dữ liệu dòng đầu
            if deriv_input_idx >= len(first_data_row) or not first_data_row[deriv_input_idx]:
                result['valid'] = False
                result['issues'].append("❌ Dòng 1: Cột 'derivative_input' trống")
                return result

            latex_input = str(first_data_row[deriv_input_idx]).strip()
            result['sample_data']['derivative_input'] = latex_input

            # Validate LaTeX format cơ bản
            if not any(keyword in latex_input for keyword in [r'\frac{d', "f'", "y'"]):
                result['warnings'].append(
                    f"⚠️ Dòng 1: LaTeX có thể không phải đạo hàm: '{latex_input}'"
                )

            print(f"✅ Pre-validation PASSED")
            print(f"   Sample: {latex_input[:50]}...")

            return result

        except Exception as e:
            result['valid'] = False
            result['issues'].append(f"❌ Lỗi validation: {str(e)}")
            return result

    # ========== STREAMING READ (FOR LARGE FILES) ==========

    def read_excel_streaming(self, file_path: str, chunksize: int = None) -> Iterator:
        """
        Đọc Excel file theo chunks với streaming (for large files)

        Args:
            file_path: Đường dẫn file Excel (.xlsx)
            chunksize: Kích thước chunk (None = auto)

        Yields:
            Iterator of (chunk_data, columns)
        """
        if chunksize is None:
            chunksize = self.estimate_optimal_chunksize(file_path)

        try:
            print(f"🚀 Streaming read với chunk size: {chunksize:,}")

            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active

            max_row = ws.max_row if hasattr(ws, 'max_row') and ws.max_row else 0

            # Header
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            columns = [str(cell) if cell is not None else f"Col_{i}"
                      for i, cell in enumerate(header_row)]

            current_row = 2
            chunk_count = 0

            while current_row <= max_row and not self.processing_cancelled:
                end_row = min(current_row + chunksize - 1, max_row)

                chunk_data = []
                for row in ws.iter_rows(min_row=current_row, max_row=end_row, values_only=True):
                    row_data = [str(cell) if cell is not None else "" for cell in row]
                    chunk_data.append(row_data)

                if chunk_data:
                    yield (chunk_data, columns)

                    del chunk_data

                    # Cleanup mỗi 10 chunks
                    if chunk_count % 10 == 0 and chunk_count > 0:
                        gc.collect()

                current_row = end_row + 1
                chunk_count += 1

            wb.close()
            print(f"✅ Streaming completed: {chunk_count} chunks")

        except Exception as e:
            raise Exception(f"Lỗi streaming: {str(e)}")

    # ========== MAIN PROCESSING (FOR LARGE FILES) ==========

    def process_large_file(
        self,
        file_path: str,
        encoding_service,
        output_path: str,
        progress_callback: Optional[Callable] = None
    ) -> Tuple[int, int, str]:
        """
        Xử lý file Excel lớn với chunked processing

        Args:
            file_path: Đường dẫn input file (.xlsx)
            encoding_service: DerivativeEncodingService instance
            output_path: Đường dẫn output file
            progress_callback: Callback(percent, current, total, errors)

        Returns:
            Tuple[success_count, error_count, output_file]
        """
        # Validate file type
        self._validate_file_type(file_path)

        self.processing_cancelled = False
        success_count = 0
        error_count = 0
        processed_count = 0
        start_time = time.time()

        # Temp file cho results
        temp_results_file = f"{output_path}.temp_results"

        try:
            print(f"🚀 CHUNKED PROCESSING: {os.path.basename(file_path)}")

            # Step 1: Check total rows
            self.total_rows = self._get_actual_total_rows(file_path)
            self._enforce_row_limit(self.total_rows)

            # Step 2: Pre-validation
            validation_result = self._validate_first_row(file_path)
            if not validation_result['valid']:
                error_msg = "❌ Validation failed!\n\n"
                for issue in validation_result['issues']:
                    error_msg += f"  • {issue}\n"
                raise ValueError(error_msg)

            # Step 3: Estimate chunk size
            chunk_size = self.estimate_optimal_chunksize(file_path)
            print(f"⚡ Chunk size: {chunk_size:,} rows")
            print(f"🎯 Target: {self.total_rows:,} rows")

            # Step 4: Process chunks
            results_buffer = []
            buffer_size = 5000
            chunk_count = 0
            last_progress_time = time.time()

            for chunk_data, columns in self.read_excel_streaming(file_path, chunk_size):
                if self.processing_cancelled:
                    break

                chunk_count += 1
                chunk_start = time.time()
                chunk_results = []

                # Find derivative_input column
                deriv_col_idx = None
                mode_col_idx = None
                for idx, col in enumerate(columns):
                    col_lower = col.strip().lower()
                    if col_lower == 'derivative_input':
                        deriv_col_idx = idx
                    elif col_lower == 'mode':
                        mode_col_idx = idx

                if deriv_col_idx is None:
                    raise Exception("Không tìm thấy cột 'derivative_input'")

                # Process rows in chunk
                for row_data in chunk_data:
                    try:
                        if self.processing_cancelled:
                            break

                        latex = row_data[deriv_col_idx].strip() if deriv_col_idx < len(row_data) else ""
                        mode = row_data[mode_col_idx].strip() if mode_col_idx and mode_col_idx < len(row_data) else "1"

                        if not latex:
                            chunk_results.append("")
                            continue

                        # Encode
                        result = encoding_service.encode_derivative(latex, mode)

                        if result['success']:
                            chunk_results.append(result['keylog'])
                            success_count += 1
                        else:
                            chunk_results.append(f"ERROR: {result.get('error', 'Unknown')}")
                            error_count += 1

                    except Exception as e:
                        chunk_results.append(f"ERROR: {str(e)}")
                        error_count += 1

                    processed_count += 1

                # Buffer results
                results_buffer.extend(chunk_results)

                if len(results_buffer) >= buffer_size:
                    self._write_results_buffer(temp_results_file, results_buffer)
                    results_buffer = []

                # Progress update
                chunk_time = time.time() - chunk_start
                chunk_speed = len(chunk_data) / chunk_time if chunk_time > 0 else 0

                current_time = time.time()
                elapsed = current_time - start_time
                avg_speed = processed_count / elapsed if elapsed > 0 else 0

                # Callback mỗi 5 giây
                if progress_callback and (current_time - last_progress_time >= 5):
                    progress_percent = (processed_count / self.total_rows) * 100
                    progress_callback(progress_percent, processed_count, self.total_rows, error_count)

                    # ETA calculation
                    remaining = self.total_rows - processed_count
                    eta_seconds = remaining / max(avg_speed, 1e-6) if remaining > 0 else 0
                    eta_str = f"{int(eta_seconds // 60):02d}:{int(eta_seconds % 60):02d}"

                    print(f"🔥 Speed: {avg_speed:.0f} rows/sec | "
                          f"Progress: {processed_count:,}/{self.total_rows:,} ({progress_percent:.1f}%) | "
                          f"ETA: {eta_str}")

                    last_progress_time = current_time

                # Memory check
                if chunk_count % 5 == 0 and self.check_memory_limit():
                    self.emergency_memory_cleanup()

                del chunk_data

            # Flush remaining buffer
            if results_buffer:
                self._write_results_buffer(temp_results_file, results_buffer)

            # Step 5: Create final Excel
            print("🔧 Creating final Excel file...")
            final_output = self._create_final_excel(file_path, temp_results_file, output_path)

            # Stats
            total_time = time.time() - start_time
            final_speed = processed_count / total_time if total_time > 0 else 0

            print(f"🏁 COMPLETED!")
            print(f"⚡ Speed: {final_speed:.0f} rows/sec")
            print(f"📊 Total: {processed_count:,} rows in {total_time:.1f}s")
            print(f"✅ Success: {success_count:,} | ❌ Errors: {error_count:,}")

            return success_count, error_count, final_output

        except Exception as e:
            raise Exception(f"Lỗi xử lý: {str(e)}")

        finally:
            # Cleanup temp file
            try:
                if os.path.exists(temp_results_file):
                    os.remove(temp_results_file)
            except:
                pass

    # ========== HELPER METHODS ==========

    def _write_results_buffer(self, temp_file: str, results: List[str]):
        """Ghi buffer results vào temp file"""
        try:
            mode = 'a' if os.path.exists(temp_file) else 'w'
            with open(temp_file, mode, encoding='utf-8', buffering=16384) as f:
                f.write('\n'.join(results) + '\n')
        except Exception as e:
            print(f"⚠️ Lỗi ghi buffer: {e}")

    def _read_temp_results(self, temp_file: str) -> List[str]:
        """Đọc results từ temp file"""
        try:
            with open(temp_file, 'r', encoding='utf-8', buffering=16384) as f:
                return f.read().strip().split('\n')
        except:
            return []

    def _create_final_excel(self, original_file: str, temp_results_file: str, output_path: str) -> str:
        """Tạo file Excel cuối cùng với keylog column"""
        try:
            print("📝 Tạo file Excel với cột keylog...")

            # Read results
            all_results = self._read_temp_results(temp_results_file)

            # Use pandas if available, otherwise openpyxl
            if PANDAS_AVAILABLE:
                return self._create_final_excel_pandas(original_file, all_results, output_path)
            else:
                return self._create_final_excel_openpyxl(original_file, all_results, output_path)

        except Exception as e:
            raise Exception(f"Lỗi tạo Excel: {str(e)}")

    def _create_final_excel_pandas(self, original_file: str, all_results: List[str], output_path: str) -> str:
        """Tạo file Excel bằng pandas (faster)"""
        try:
            import pandas as pd

            # Read original file
            original_df = pd.read_excel(original_file, dtype=str, keep_default_na=False, engine='openpyxl')

            # Trim results to match df length
            results_to_add = all_results[:len(original_df)]
            if len(results_to_add) < len(original_df):
                results_to_add.extend([''] * (len(original_df) - len(results_to_add)))

            # Add/update keylog column
            if 'keylog' in [c.strip().lower() for c in original_df.columns]:
                exact_name = next((c for c in original_df.columns if c.strip().lower() == 'keylog'), 'keylog')
                original_df[exact_name] = results_to_add
            else:
                original_df['keylog'] = results_to_add

            # Write to Excel with formatting
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                original_df.to_excel(writer, sheet_name='Results', index=False)
                worksheet = writer.sheets['Results']

                # Find keylog column
                header_cells = next(worksheet.iter_rows(min_row=1, max_row=1))
                keylog_col_idx = None
                for idx, cell in enumerate(header_cells, start=1):
                    if str(cell.value).strip().lower() == 'keylog':
                        keylog_col_idx = idx
                        break

                # Apply font to keylog column
                if keylog_col_idx:
                    keylog_font = Font(name="Flexio Fx799VN", size=11, bold=True, color="000000")
                    max_row = worksheet.max_row
                    for row in range(1, max_row + 1):
                        worksheet.cell(row=row, column=keylog_col_idx).font = keylog_font

                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = min(len(str(cell.value)), 40)
                        except:
                            pass
                    worksheet.column_dimensions[column_letter].width = max(max_length + 2, 10)

            print(f"✅ File Excel đã tạo: {output_path}")
            return output_path

        except Exception as e:
            # Fallback to openpyxl
            return self._create_final_excel_openpyxl(original_file, all_results, output_path)

    def _create_final_excel_openpyxl(self, original_file: str, all_results: List[str], output_path: str) -> str:
        """Tạo file Excel bằng openpyxl (fallback)"""
        try:
            wb = openpyxl.load_workbook(original_file, data_only=True)
            ws = wb.active

            # Find or add keylog column
            keylog_col_idx = None
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            for col_idx, val in enumerate(header_row, 1):
                if val and str(val).strip().lower() == 'keylog':
                    keylog_col_idx = col_idx
                    break

            if keylog_col_idx is None:
                keylog_col_idx = ws.max_column + 1

            keylog_col_letter = get_column_letter(keylog_col_idx)
            ws[f'{keylog_col_letter}1'].value = 'keylog'

            # Write results
            for idx, result in enumerate(all_results):
                excel_row = idx + 2
                if excel_row > ws.max_row:
                    break
                ws[f'{keylog_col_letter}{excel_row}'].value = result

            # Apply font
            keylog_font = Font(name="Flexio Fx799VN", size=11, bold=True, color="000000")
            for row in range(1, ws.max_row + 1):
                ws[f'{keylog_col_letter}{row}'].font = keylog_font

            wb.save(output_path)
            wb.close()

            print(f"✅ File Excel đã tạo: {output_path}")
            return output_path

        except Exception as e:
            raise Exception(f"Lỗi tạo Excel (openpyxl): {str(e)}")

    # ========== PUBLIC API (FOR NORMAL FILES) ==========

    def read_excel_file(self, file_path: str) -> Tuple[bool, List[Tuple[str, str]], str]:
        """
        Đọc file Excel và trả về rows (cho file nhỏ)
        ⚠️ CHỈ HỖ TRỢ XLSX
        """
        try:
            # Validate file type
            self._validate_file_type(file_path)

            self.file_path = file_path
            self.file_size_mb = os.path.getsize(file_path) / (1024 * 1024)

            rows = []
            deriv_input_col_idx = -1
            mode_col_idx = -1

            wb = openpyxl.load_workbook(file_path, data_only=False)
            ws = wb.active

            for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                if row_idx == 1:
                    # Header row
                    for col_idx, header in enumerate(row):
                        if header:
                            header_lower = str(header).strip().lower()
                            if header_lower == 'derivative_input':
                                deriv_input_col_idx = col_idx
                            elif header_lower == 'mode':
                                mode_col_idx = col_idx

                    if deriv_input_col_idx == -1:
                        wb.close()
                        return False, [], "❌ Không tìm thấy cột 'derivative_input'"
                else:
                    # Data rows
                    if row and deriv_input_col_idx < len(row) and row[deriv_input_col_idx]:
                        deriv_input = str(row[deriv_input_col_idx]).strip()
                        mode = "1"  # Default

                        if mode_col_idx != -1 and mode_col_idx < len(row) and row[mode_col_idx]:
                            _mode = str(row[mode_col_idx]).strip()
                            if _mode:
                                mode = _mode

                        rows.append((deriv_input, mode))

            wb.close()
            return True, rows, ""

        except ValueError as ve:
            # File type error
            return False, [], str(ve)
        except Exception as e:
            return False, [], f"❌ Lỗi đọc file: {str(e)}"

    def export_results(self, batch_results: List[Dict], use_csv: Optional[bool] = None) -> Tuple[bool, str, str]:
        """
        Export results vào file Excel (cho file nhỏ)
        """
        try:
            if not self.file_path:
                return False, "", "❌ Chưa chọn file"

            wb = openpyxl.load_workbook(self.file_path, data_only=False)
            ws = wb.active

            # Find/add keylog column
            keylog_col_idx = None
            header_row = [cell.value for cell in ws[1]]
            for col_idx, val in enumerate(header_row, 1):
                if val and str(val).strip().lower() == 'keylog':
                    keylog_col_idx = col_idx
                    break

            if keylog_col_idx is None:
                keylog_col_idx = ws.max_column + 1
                keylog_col_letter = get_column_letter(keylog_col_idx)
                ws[f'{keylog_col_letter}1'].value = 'keylog'
            else:
                keylog_col_letter = get_column_letter(keylog_col_idx)

            # Write results with chunking
            chunk_size = 1000
            total_rows = len(batch_results)
            num_chunks = math.ceil(total_rows / chunk_size)

            for chunk_i in range(num_chunks):
                start_idx = chunk_i * chunk_size
                end_idx = min(start_idx + chunk_size, total_rows)
                for idx in range(start_idx, end_idx):
                    excel_row = idx + 2
                    ws[f'{keylog_col_letter}{excel_row}'].value = batch_results[idx]['keylog']

            output_file = self._get_output_file_path('.xlsx')
            wb.save(output_file)
            wb.close()

            return True, output_file, f"✅ File đã lưu: {output_file}"

        except Exception as e:
            return False, "", f"❌ Lỗi export: {str(e)}"

    def _get_output_file_path(self, extension: str) -> str:
        """Tạo đường dẫn file output với timestamp và số dòng"""
        file_name = os.path.basename(self.file_path)
        file_name_without_ext = os.path.splitext(file_name)[0]
        output_dir = os.path.dirname(self.file_path)
        timestamp = datetime.now().strftime("%Y%m%d")

        # Đếm số dòng dữ liệu (không tính header)
        try:
            import openpyxl
            wb = openpyxl.load_workbook(self.file_path, read_only=True, data_only=True)
            ws = wb.active
            row_count = max(0, ws.max_row - 1) if hasattr(ws, 'max_row') and ws.max_row else 0
            wb.close()
        except Exception as e:
            print(f"Không thể đếm số dòng: {e}, sử dụng pandas")
            try:
                import pandas as pd
                df = pd.read_excel(self.file_path, nrows=0)  # Chỉ đọc header để kiểm tra
                df_full = pd.read_excel(self.file_path)
                row_count = len(df_full)
            except Exception:
                row_count = 0

        # Format: filename_encoded_YYYYMMDD_XXXXrows.xlsx
        return os.path.join(
            output_dir,
            f"{file_name_without_ext}_encoded_{timestamp}_{row_count}rows{extension}"
        )

    def get_file_info(self) -> Dict:
        """Lấy thông tin file"""
        return {
            'path': self.file_path,
            'size_mb': round(self.file_size_mb, 2),
            'total_rows': self.total_rows,
            'threshold_mb': self.FILE_SIZE_THRESHOLD_MB,
            'max_rows_allowed': self.MAX_ROWS_ALLOWED,
            'is_large_file': self.file_size_mb > self.FILE_SIZE_THRESHOLD_MB,
            'supported_format': 'XLSX only'
        }
