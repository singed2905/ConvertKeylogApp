import os
import csv
import openpyxl
from openpyxl.utils import get_column_letter
from typing import List, Tuple, Dict, Optional
from datetime import datetime
import math


class ExcelService:
    """Service xử lý import/export Excel cho Integral Mode - chunked export"""

    FILE_SIZE_THRESHOLD = 50

    def __init__(self):
        self.file_path = None
        self.file_size_mb = 0
        self.use_csv = False  # Luôn xuất Excel kể cả file lớn

    def read_excel_file(self, file_path: str) -> Tuple[bool, List[Tuple[str, str]], str]:
        """Đọc file Excel/CSV, luôn ưu tiên cột int_input"""
        try:
            self.file_path = file_path
            self.file_size_mb = os.path.getsize(file_path) / (1024 * 1024)

            rows = []
            int_input_col_idx = -1
            mode_col_idx = -1

            # Đọc file (CSV hoặc Excel)
            if file_path.endswith('.csv'):
                with open(file_path, 'r', encoding='utf-8') as f:
                    reader = csv.reader(f)
                    for row_idx, row in enumerate(reader):
                        if row_idx == 0:
                            for col_idx, header in enumerate(row):
                                header_lower = header.strip().lower()
                                if header_lower == 'int_input':
                                    int_input_col_idx = col_idx
                                elif header_lower == 'mode':
                                    mode_col_idx = col_idx
                            if int_input_col_idx == -1:
                                return False, [], "❌ Không tìm thấy cột 'int_input'"
                        else:
                            if len(row) > int_input_col_idx and row[int_input_col_idx]:
                                int_input = row[int_input_col_idx].strip()
                                mode = "3"
                                if mode_col_idx != -1 and len(row) > mode_col_idx:
                                    _mode = row[mode_col_idx].strip()
                                    if _mode:
                                        mode = _mode
                                rows.append((int_input, mode))

            else:  # Excel file
                wb = openpyxl.load_workbook(file_path, data_only=False)
                ws = wb.active
                for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                    if row_idx == 1:
                        for col_idx, header in enumerate(row):
                            if header:
                                header_lower = str(header).strip().lower()
                                if header_lower == 'int_input':
                                    int_input_col_idx = col_idx
                                elif header_lower == 'mode':
                                    mode_col_idx = col_idx
                        if int_input_col_idx == -1:
                            wb.close()
                            return False, [], "❌ Không tìm thấy cột 'int_input'"
                    else:
                        if row and int_input_col_idx < len(row) and row[int_input_col_idx]:
                            int_input = str(row[int_input_col_idx]).strip()
                            mode = "3"
                            if mode_col_idx != -1 and mode_col_idx < len(row) and row[mode_col_idx]:
                                _mode = str(row[mode_col_idx]).strip()
                                if _mode:
                                    mode = _mode
                            rows.append((int_input, mode))
                wb.close()

            # Chỉ xuất Excel dù file lớn
            self.use_csv = False
            return True, rows, ""
        except Exception as e:
            return False, [], f"❌ Lỗi đọc file: {str(e)}"

    def export_results(self, batch_results: List[Dict], use_csv: Optional[bool] = None) -> Tuple[bool, str, str]:
        """Export kết quả keylog vào file (dùng chunk, luôn xuất file Excel)"""
        try:
            if not self.file_path:
                return False, "", "❌ Chưa chọn file"
            # Luôn export Excel
            return self._export_as_excel(batch_results)
        except Exception as e:
            return False, "", f"❌ Lỗi export: {str(e)}"

    def _export_as_excel(self, batch_results: List[Dict]) -> Tuple[bool, str, str]:
        """Export as Excel file với xử lý theo chunk (ổn định cho file lớn)"""
        try:
            wb = openpyxl.load_workbook(self.file_path, data_only=False)
            ws = wb.active

            # Tìm hoặc thêm cột 'keylog'
            keylog_col_idx = None
            header_row = [cell.value for cell in ws[1]]
            for col_idx, val in enumerate(header_row, 1):
                if val and str(val).strip().lower() == 'keylog':
                    keylog_col_idx = col_idx
                    break

            if keylog_col_idx is None:
                keylog_col_idx = ws.max_column + 1
                keylog_col_letter = get_column_letter(keylog_col_idx)
                ws[f'{keylog_col_letter}1'].value = 'keylog'
            else:
                keylog_col_letter = get_column_letter(keylog_col_idx)

            chunk_size = 500  # chunk kích thước, có thể tăng lên nếu máy mạnh
            total_rows = len(batch_results)
            num_chunks = math.ceil(total_rows / chunk_size)

            for chunk_i in range(num_chunks):
                start_idx = chunk_i * chunk_size
                end_idx = min(start_idx + chunk_size, total_rows)
                for idx in range(start_idx, end_idx):
                    excel_row = idx + 2  # offset 2 vì header là dòng 1
                    ws[f'{keylog_col_letter}{excel_row}'].value = batch_results[idx]['keylog']
                # Nhận xét: Nếu file cực lớn, có thể gọi wb.save(tênfile_tam) từng chunk

            output_file = self._get_output_file_path('.xlsx')
            wb.save(output_file)
            wb.close()

            return True, output_file, f"✅ File Excel đã lưu: {output_file}"

        except Exception as e:
            return False, "", f"❌ Lỗi export Excel chunk: {str(e)}"

    def _get_output_file_path(self, extension: str) -> str:
        """Tạo đường dẫn file output với timestamp"""
        file_name = os.path.basename(self.file_path)
        file_name_without_ext = os.path.splitext(file_name)[0]
        output_dir = os.path.dirname(self.file_path)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return os.path.join(output_dir, f"{file_name_without_ext}_encoded_{timestamp}{extension}")

    def get_file_info(self) -> Dict:
        """Lấy thông tin file"""
        return {
            'path': self.file_path,
            'size_mb': round(self.file_size_mb, 2),
            'use_csv': self.use_csv,
            'threshold_mb': self.FILE_SIZE_THRESHOLD
        }
